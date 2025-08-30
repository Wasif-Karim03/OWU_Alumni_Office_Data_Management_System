#!/usr/bin/env python3
"""
Create Ohio Wesleyan University Event Financial Data Template
This creates an Excel file with the proper structure for event financial tracking
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
import openpyxl.styles
import openpyxl.utils

def create_owu_template():
    """Create a proper Excel template for OWU event financial data."""
    
    # Financial categories (left column)
    financial_categories = [
        'Event Income',
        'All Incurred Expenses (food, beverage, rental etc.)',
        'Underwritten',
        'Profit/Loss'
    ]
    
    # Sample events with dates (top row)
    events = [
        '07/27/2024 OWU Football Golf Outing',
        '08/02/2024 OWU Clippers Home Game',
        '08/15/2024 OWU Legacy Reception',
        '08/21/2024 OWU Senior Class Welcome Back',
        '08/25/2024 OWU Monnett Club Event',
        '09/11/2024 OWU Near You - Toledo',
        '09/11/2024 OWU Near You - Denver',
        '09/12/2024 OWU Near You - Tucson',
        '09/12/2024 OWU Near You - Cleveland',
        '09/12/2024 OWU Near You - Atlanta',
        '09/12/2024 OWU Near You - Chicago',
        '09/13/2024 OWU Near You - Cincinnati',
        '09/15/2024 OWU Near You - Charlotte',
        '09/16/2024 OWU Near You - Ann Arbor',
        '09/17/2024 OWU Near You - Newport Beach',
        '09/18/2024 OWU Near You - Seattle',
        '09/18/2024 WOW Speaker Series - Virtual',
        '09/19/2024 OWU Near You - Columbus',
        '09/25/2024 OWU Monnett Club',
        '09/27/2024 Women\'s Basketball',
        '10/03/2024 Emerging Leaders with Alumni Board',
        '10/04-05/2024 Homecoming & Family Weekend',
        '10/04/2024 Golf Outing',
        '10/05/2024 Crew Game',
        '10/26/2024 Men\'s/Women\'s Soccer Weekend',
        '11/08-09/2024 Women of Wesleyan',
        '11/09/2024 Delta Zeta Dinner',
        '12/04/2024 Holiday Party - Columbus',
        '12/05/2024 Holiday Party - NYC',
        '12/10/2024 Holiday Party - Denver',
        '12/11/2024 Holiday Party - Cleveland',
        '12/12/2024 Holiday Party - DC',
        '01/11/2025 Red and Black Track and Field Alumni Meet',
        '01/08/2025 WOW Strengths 2.0 (Virtual)',
        '01/20/2025 MLK Breakfast',
        '02/14/2025 Naples Luncheon',
        '03/24/2025 Ask a Bishop/LLI Deydre Teyhen (Virtual)',
        '03/27-29/2025 Melvin Van Peebles Symposium',
        '03/31/2025 Ask a Bishop/LLI Diane Petersen (Virtual)',
        '04/05/2025 FIJI Pig Dinner',
        '04/05/2025 Monnett at Scioto CC',
        '04/07/2025 Ask a Bishop/LLI Kara Trott (Virtual)',
        '04/10/2025 Senior Class Almost Alumni Event with Alumni Board',
        '04/11-12/2025 Jay Martin Retirement Celebration',
        '04/14/2025 Ask A Bishop/LLI Eric Buer (Virtual)',
        '05/08/2025 Senior Class Final Lap',
        '05/16-18/2025 Reunion Weekend',
        '06/07/2025 OWU at the Zoo'
    ]
    
    # Create the data structure
    data = {}
    
    # Add financial categories as the first column
    data['Financial Categories'] = financial_categories
    
    # Add sample financial data for each event
    for event in events:
        # Generate realistic sample data
        if 'Golf' in event:
            data[event] = [350, 500, 0, -150]  # Income, Expenses, Underwritten, Profit/Loss
        elif 'Homecoming' in event:
            data[event] = [10073, 28633, 0, -18560]
        elif 'Holiday Party' in event:
            data[event] = [350, 1280, 0, -930]
        elif 'OWU Near You' in event:
            data[event] = [0, 14, 0, -14]
        elif 'Legacy Reception' in event:
            data[event] = [0, 501, 0, -501]
        elif 'Monnett Club' in event:
            data[event] = [0, 275, 0, -275]
        else:
            # Default sample data
            data[event] = [0, 100, 0, -100]
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Save to Excel with proper formatting
    filename = 'OWU_Event_Financial_Template.xlsx'
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Event Financial Data', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Event Financial Data']
        
        # Format the header row (event names)
        for col in range(1, len(events) + 2):  # +2 because first col is categories
            cell = worksheet.cell(row=1, column=col)
            cell.font = cell.font.copy(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='D00000', end_color='D00000', fill_type='solid')
            cell.font = cell.font.copy(color='FFFFFF')
        
        # Format the first column (financial categories)
        for row in range(1, len(financial_categories) + 2):  # +2 because first row is header
            cell = worksheet.cell(row=row, column=1)
            cell.font = cell.font.copy(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
            cell.font = cell.font.copy(color='FFFFFF')
        
        # Format financial data cells
        for row in range(2, len(financial_categories) + 2):
            for col in range(2, len(events) + 2):
                cell = worksheet.cell(row=row, column=col)
                if row == 4:  # Profit/Loss row
                    if cell.value and cell.value < 0:
                        cell.font = cell.font.copy(color='D00000')  # Red for losses
                    elif cell.value and cell.value > 0:
                        cell.font = cell.font.copy(color='27AE60')  # Green for profits
        
        # Adjust column widths
        for col in range(1, len(events) + 2):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
        
        # Freeze the first row and first column
        worksheet.freeze_panes = 'B2'
    
    print(f"✅ OWU Event Financial Template created: {filename}")
    print(f"📊 Template includes {len(events)} events and {len(financial_categories)} financial categories")
    print(f"🎯 Perfect structure for your Ohio Wesleyan University event tracking!")
    
    return filename

if __name__ == "__main__":
    try:
        create_owu_template()
    except ImportError as e:
        print("❌ Error: openpyxl not installed. Installing...")
        import subprocess
        subprocess.run(['pip', 'install', 'openpyxl'])
        print("✅ openpyxl installed. Running template creation...")
        create_owu_template()
